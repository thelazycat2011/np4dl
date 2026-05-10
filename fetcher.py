import requests
import openpyxl
import json
from io import BytesIO
SHEET_ID = '1qGjbhVHoIWlxgFE5AV1caHnTyvqxkFWQEBbs9G96ZVI'
LIMIT = 200
url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
response = requests.get(url)
wb = openpyxl.load_workbook(filename=BytesIO(response.content), data_only=True)

def builddemonlist(wb):
    sheet = wb.worksheets[1]
    listbuilder = []
    counter = 0
    f = True
    for row in sheet.iter_rows():
        counter += 1
        data = {
            "id": 0,
            "name": "",
            "author": "",
            "creators": [],
            "verifier": "",
            "verification": "",
            "percentToQualify": "100",
            "password": "No Password",
            "records": []
        }
        if counter < 7: continue
        if counter > LIMIT + 6: break
        row_data = []
        for id, cell in enumerate(row):
            link = cell.hyperlink.target if cell.hyperlink else None
            if id == 1 and cell.value == "":
                f = False
                break
            elif id == 1:
                listbuilder.append(f"level{counter - 6}")
                data["name"] = cell.value
            elif id == 2: data["author"] = cell.value
            elif id == 3: 
                try:
                    data["id"] = int(cell.value)
                except:
                    f = False
                    listbuilder.pop(-1)
                    break
            elif id == 6: data["verifier"] = cell.value
            elif id == 7: data["verification"] = link
            if not f: break
        if f:
            with open(f"data/level{counter - 6}.json", "w") as file: file.write(json.dumps(data))
        else: break
    with open(f"data/_list.json", "w") as file: file.write(json.dumps(listbuilder))

def populaterecords(wb):
    sheet = wb.worksheets[2]
    for idx, row in enumerate(sheet.iter_rows()):
        if idx < 2: continue
        if idx > LIMIT + 1: break
        data = {}
        with open(f"data/level{idx - 1}.json", "r") as file: data = json.loads(file.read())
        for id, cell in enumerate(row):
            template = {
                "user": "",
                "link": "",
                "percent": 100,
                "hz": 60
            }
            link = cell.hyperlink.target if cell.hyperlink else None
            if id > 3 and cell.value == None: break
            elif id > 3:
                template["user"] = cell.value
                template["link"] = link
                data["records"].append(template)
        with open(f"data/level{idx - 1}.json", "w") as file: file.write(json.dumps(data))

builddemonlist(wb)
populaterecords(wb)